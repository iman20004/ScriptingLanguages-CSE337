########## IMAN ALI ##########
########## imaali ##############
########## 112204305 #############

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
import os
import re
from openpyxl.worksheet.worksheet import Worksheet


class ExcelFileManager:
    def __init__(self, filename):
        if os.path.isdir(filename):
            raise ValueError()
        self.filename = filename

    def write_sheet(self, data, sheetname, wb=None, save=True):
        if not wb:
            wb = Workbook()
            sheet = wb.create_sheet(sheetname)
            curr_row = 1
        else:
            if sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                curr_row = sheet.max_row + 1
            else:
                sheet = wb.create_sheet(sheetname)
                curr_row = 1

        for tup in data:
            # Check if int, or float or string is a number
            if type(tup[1]) == str:
                if not tup[1].isnumeric():
                    if not tup[1].replace('.', '', 1).isdigit():
                        save = False
                        break

            # Check if student already in the work sheet
            skip = False
            row_num = 1
            col_num = 1
            for row in sheet.iter_rows():
                if row[0].value == tup[0]:  # Name of student found
                    for _ in row:  # Find col number of next empty cell in that row
                        col_num += 1
                    skip = True
                    break
                row_num += 1

            if skip:
                sheet.cell(row=row_num, column=col_num).value = float(tup[1])
                continue

            sheet.cell(row=curr_row, column=1).value = tup[0]
            sheet.cell(row=curr_row, column=2).value = float(tup[1])
            curr_row += 1

        # either save and return 0 or return -1
        if save:
            wb.save(self.filename)
            return 0
        else:
            return -1

    def write_sheets(self, data):
        wb = Workbook()
        wb.save(self.filename)
        for tup in data:
            if self.write_sheet([tup], tup[2], wb) == -1:
                os.remove(self.filename)
                return -1
        return 0

    def get_sheet_ave(self, sheetname):
        avg_list = []
        try:
            wb = openpyxl.load_workbook(self.filename)
            sheet = wb[sheetname]
        except (KeyError, InvalidFileException, FileNotFoundError) as e:
            return avg_list

        for row in sheet.rows:
            total_score = 0
            num_tests = 0
            for cell in row:
                if type(cell.value) == str or not cell.value:
                    continue
                total_score += cell.value
                num_tests += 1

            if num_tests == 0:
                return []
            avg_list.append((row[0].value, total_score / num_tests))

        return avg_list

    def get_workbook_ave(self, pattern=None):
        mean_list = []
        try:
            wb = openpyxl.load_workbook(self.filename)
        except (InvalidFileException, FileNotFoundError) as e:
            return mean_list

        if pattern:
            sheet_list = []
            for s in wb.sheetnames:
                if re.search(pattern, s):
                    sheet_list.append(s)
        else:
            sheet_list = wb.sheetnames
        name_score_num = {}

        for sheetname in sheet_list:
            tup = self.get_sheet_ave(sheetname)
            if not tup:
                continue

            for tu in tup:
                if tu[0] in name_score_num:
                    name_score_num[tu[0]] = (name_score_num[tu[0]][0]+tu[1], name_score_num[tu[0]][1]+1)
                    print(name_score_num[tu[0]])
                else:
                    name_score_num[tu[0]] = (tu[1], 1)

        for key in name_score_num:
            mean_list.append((key, name_score_num[key][0]/name_score_num[key][1]))

        return mean_list
