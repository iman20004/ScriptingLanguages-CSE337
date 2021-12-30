import unittest
import os
import shutil
import openpyxl
import sys
sys.path.insert(1,os.getcwd())
from excel_file_mgr import ExcelFileManager

class ExcelFileManagerTest(unittest.TestCase):

    def setUp(self):
        if os.path.isdir('tmp-excel'):
            shutil.rmtree('tmp-excel')
        else:
            os.mkdir('tmp-excel')

    def tearDown(self):
        if os.path.isdir('tmp-excel'):
            shutil.rmtree('tmp-excel')

    def test_excel_file_mgr_raise_value_error(self):
        with self.assertRaises(ValueError):
            ExcelFileManager('tmp-excel')

    def test_excel_file_mgr_set_filename(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        self.assertEqual(os.path.join('tmp-excel','sample.xlsx'), f.filename)


    def test_excel_file_mgr_write_sheet_numeric_scores(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kebla', 98.2), ('Jhamela', 85), ('Kebla', 98.2)]
        f.write_sheet(data, 'CSE337')
        wb = openpyxl.load_workbook(f.filename)
        sheet = wb['CSE337']
        self.assertEqual('Kebla', sheet.cell(row=1,column=1).value)
        self.assertEqual(98.2, sheet.cell(row=1,column=2).value)
        self.assertEqual(98.2, sheet.cell(row=1,column=3).value)
        self.assertEqual('Jhamela', sheet.cell(row=2,column=1).value)
        self.assertEqual(85.0, sheet.cell(row=2,column=2).value)

    def test_excel_file_mgr_write_sheet_string_scores(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kebla', '98.2'), ('Jhamela', '85'), ('Kebla', '98.2')]
        f.write_sheet(data, 'ISE337')
        wb = openpyxl.load_workbook(f.filename)
        sheet = wb['ISE337']
        self.assertEqual('Kebla', sheet.cell(row=1,column=1).value)
        self.assertEqual(98.2, sheet.cell(row=1,column=2).value)
        self.assertEqual(98.2, sheet.cell(row=1,column=3).value)
        self.assertEqual('Jhamela', sheet.cell(row=2,column=1).value)
        self.assertEqual(85.0, sheet.cell(row=2,column=2).value)

    def test_excel_file_mgr_write_sheet_string_numeric(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kebla', 98.2), ('Keramoti', '85'), ('Kebla', '98.2')]
        self.assertEqual(0, f.write_sheet(data, 'ISE337'))

    def test_excel_file_mgr_write_sheet_invalid_score_1(self):
        f = ExcelFileManager(os.path.join('someDir','sample.xlsx'))
        data = [('Kebla', 98.2), ('Keramoti', 'ten'), ('Kebla', '98.2')]
        self.assertEqual(-1, f.write_sheet(data, 'CSE337'))

    def test_excel_file_mgr_write_sheet_invalid_score_2(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kebla', 98.2), ('Keramoti', 'ten'), ('Kebla', '98.2')]
        self.assertEqual(-1, f.write_sheet(data, 'CSE337'))

    def test_excel_file_mgr_write_multi_sheets_1(self):
        data = [('Jonny', 91, 'CSE337'),('Kebla', 57, 'ISE337'),('Kebla', 79, 'CSE337'), ('Bombol', 99, 'ISE337')]
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        f.write_sheets(data)
        wb = openpyxl.load_workbook(f.filename)
        sheet1 = wb['CSE337']
        sheet2 = wb['ISE337']
        self.assertEqual('Jonny', sheet1.cell(row=1,column=1).value)
        self.assertEqual(91, sheet1.cell(row=1,column=2).value)
        self.assertEqual('Kebla', sheet1.cell(row=2,column=1).value)
        self.assertEqual(79, sheet1.cell(row=2,column=2).value)

        self.assertEqual('Kebla', sheet2.cell(row=1,column=1).value)
        self.assertEqual(57, sheet2.cell(row=1,column=2).value)
        self.assertEqual('Bombol', sheet2.cell(row=2,column=1).value)
        self.assertEqual(99, sheet2.cell(row=2,column=2).value)

    def test_excel_file_mgr_write_multi_sheets_2(self):
        data = [('Jonny', 91, 'Sheet1'),('Joma', '57.5', 'ISE337'),('Kebla', 79.9, 'CSE337'), ('Bombol', 99, 'ISE337')]
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        f.write_sheets(data)
        wb = openpyxl.load_workbook(f.filename)
        sheet1 = wb['CSE337']
        sheet2 = wb['ISE337']
        sheet3 = wb['Sheet1']
        self.assertEqual('Jonny', sheet3.cell(row=1,column=1).value)
        self.assertEqual(91, sheet3.cell(row=1,column=2).value)

        self.assertEqual('Kebla', sheet1.cell(row=1,column=1).value)
        self.assertEqual(79.9, sheet1.cell(row=1,column=2).value)

        self.assertEqual('Joma', sheet2.cell(row=1,column=1).value)
        self.assertEqual(57.5, sheet2.cell(row=1,column=2).value)
        self.assertEqual('Bombol', sheet2.cell(row=2,column=1).value)
        self.assertEqual(99, sheet2.cell(row=2,column=2).value)

    def test_excel_file_mgr_write_sheets_reject_string_score(self):
        data = [('Jonny', 91, 'Sheet1'),('Joma', '57.5', 'ISE337'),('Kebla', 79.9, 'CSE337'), ('Bombol', 'ninety nine', 'ISE337')]
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        self.assertEqual(-1, f.write_sheets(data))
        self.assertFalse(os.path.exists(os.path.join('tmp-excel','sample.xlsx')))

    def test_excel_file_mgr_get_sheet_ave_simple(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kebla', 98.2), ('Jhamela', 85), ('Kebla', 98.2)]
        f.write_sheet(data, 'CISE337')
        self.assertIn(('Kebla', 98.2), f.get_sheet_ave('CISE337'))
        self.assertIn(('Jhamela', 85), f.get_sheet_ave('CISE337'))

    def test_excel_file_mgr_get_sheet_ave_complex(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kumir', '91.2'), ('Pagol', 43.00),
        ('Potol', 73), ('Posto', 86.2),
        ('Kumir', 90.0), ('Pagol', 59),
        ('Pagol', '60.0'), ('Potol', 75),
        ('Potol', '93'), ('Posto', 22)]
        f.write_sheet(data, 'CISE337')
        self.assertIn(('Kumir', (91.2+90.0)/2), f.get_sheet_ave('CISE337'))
        self.assertIn(('Pagol', (43.00+59+60.0)/3), f.get_sheet_ave('CISE337'))
        self.assertIn(('Potol', (93+75+73)/3), f.get_sheet_ave('CISE337'))
        self.assertIn(('Posto', (86.2+22)/2), f.get_sheet_ave('CISE337'))

    def test_excel_file_mgr_get_sheet_ave_symmetry(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kumir', '91.2'), ('Pagol', 43.00), ('Kumir', 44),
        ('Potol', 73), ('Posto', 86.2), ('Pagol', 89), ('Kumir', 68),
        ('Kumir', 90.0), ('Pagol', 59),
        ('Pagol', '60.0'), ('Potol', 75), ('Potol', 75), ('Posto', 72),
        ('Potol', '93'), ('Posto', 22), ('Posto', '100')]
        f.write_sheet(data, 'CISE337')
        self.assertIn(('Kumir', (91.2+44+68+90.0)/4), f.get_sheet_ave('CISE337'))
        self.assertIn(('Pagol', (43.00+89+59+60.0)/4), f.get_sheet_ave('CISE337'))
        self.assertIn(('Potol', (93+75+73+75)/4), f.get_sheet_ave('CISE337'))
        self.assertIn(('Posto', (86.2+22+72+100)/4), f.get_sheet_ave('CISE337'))

    def test_excel_file_mgr_get_sheet_no_workbook(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        self.assertEqual([], f.get_sheet_ave('CISE337'))

    def test_excel_file_mgr_get_sheet_no_sheet(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kumir', '91.2'), ('Pagol', 43.00), ('Kumir', 44),
        ('Potol', 73), ('Posto', 86.2), ('Pagol', 89), ('Kumir', 68),
        ('Kumir', 90.0), ('Pagol', 59),
        ('Pagol', '60.0'), ('Potol', 75), ('Potol', 75), ('Posto', 72),
        ('Potol', '93'), ('Posto', 22), ('Posto', '100')]
        f.write_sheet(data, 'CISE337')
        self.assertEqual([], f.get_sheet_ave('Sheet_A'))

    def test_excel_file_mgr_get_sheet_bad_file(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.txt'))
        data = [('Kumir', '91.2'), ('Pagol', 43.00), ('Kumir', 44),
        ('Potol', 73), ('Posto', 86.2), ('Pagol', 89), ('Kumir', 68),
        ('Kumir', 90.0), ('Pagol', 59),
        ('Pagol', '60.0'), ('Potol', 75), ('Potol', 75), ('Posto', 72),
        ('Potol', '93'), ('Posto', 22), ('Posto', '100')]
        f.write_sheet(data, 'CISE337')
        self.assertEqual([], f.get_sheet_ave('CISE337'))

    def test_excel_file_mgr_get_workbook_ave_simple(self):
        data = [('Jonny', 91, 'CSE337'),('Kebla', 57, 'ISE337'),('Kebla', 79, 'CSE337'), ('Bombol', 99, 'ISE337')]
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        f.write_sheets(data)
        self.assertIn(('Jonny', 91), f.get_workbook_ave('^(C|I)SE337$'))
        self.assertIn(('Kebla', (57+79)/2), f.get_workbook_ave('^C|ISE337$'))
        self.assertIn(('Bombol', 99), f.get_workbook_ave('^(C|I)SE337$'))

    def test_excel_file_mgr_get_workbook_ave_simple_1(self):
        data = [('Kebla', 20, 'CSE337'),('Kebla', 10, 'CSE337'),('Kebla', 15, 'ISE337'), ('kebla', 10, 'ISE337'), ('kebla', 20, 'ISE337')]
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        f.write_sheets(data)
        self.assertIn(('Kebla', 15), f.get_workbook_ave('^(C|I)SE337$'))

    def test_excel_file_mgr_get_workbook_ave_missing_pattern(self):
        data = [('Jonny', 91, 'CSE337'),('Kebla', 57, 'ISE337'),('Kebla', 79, 'CSE337'), ('Bombol', 99, 'ISE337')]
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        f.write_sheets(data)
        self.assertIn(('Jonny', 91), f.get_workbook_ave())
        self.assertIn(('Kebla', (57+79)/2), f.get_workbook_ave())
        self.assertIn(('Bombol', 99), f.get_workbook_ave())

    def test_excel_file_mgr_get_workbook_ave_complex(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kumir', '91.2', 'CSE337'), ('Pagol', 43.00, 'ISE337'),
        ('Potol', 73, 'ISE337'), ('Posto', 86.2, 'ISE337'),
        ('Kumir', 90.0, 'CS337'), ('Pagol', 59, 'CSE337'),
        ('Pagol', '60.0', 'CSE337'), ('Potol', 75, 'ISE337'),
        ('Potol', '93', 'ISE337'), ('Posto', 22, 'SE337')]
        f.write_sheets(data)
        self.assertIn(('Kumir', 91.2), f.get_workbook_ave('^(C|I)SE337$'))
        self.assertIn(('Pagol', (43.00 + (59 + 60.0)/2)/2), f.get_workbook_ave('^(C|I)SE337$'))
        self.assertIn(('Potol', (73 + 75 + 93)/3), f.get_workbook_ave('^(C|I)SE337$'))
        self.assertIn(('Posto', 86.2), f.get_workbook_ave('^(C|I)SE337$'))

    def test_excel_file_mgr_get_workbook_ave_complex_no_pattern(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kumir', '91.2', 'CSE337'), ('Pagol', 43.00, 'ISE337'),
        ('Potol', 73, 'ISE337'), ('Posto', 86.2, 'ISE337'),
        ('Kumir', 90.0, 'CS337'), ('Pagol', 59, 'CSE337'),
        ('Pagol', '60.0', 'CSE337'), ('Potol', 75, 'CSE337'),
        ('Potol', '93', 'ISE337'), ('Posto', 22, 'SE337')]
        f.write_sheets(data)
        self.assertIn(('Kumir', (91.2 + 90.0)/2), f.get_workbook_ave())
        self.assertIn(('Pagol', (43.00 + (59 + 60.0)/2)/2), f.get_workbook_ave())
        self.assertIn(('Potol', ((73 + 93)/2 + 75)/2), f.get_workbook_ave())
        self.assertIn(('Posto', (86.2+22)/2), f.get_workbook_ave())

    def test_excel_file_mgr_get_workbook_ave_complex_no_match(self):
        f = ExcelFileManager(os.path.join('tmp-excel','sample.xlsx'))
        data = [('Kumir', '91.2', 'CSE337'), ('Pagol', 43.00, 'ISE337'),
        ('Potol', 73, 'ISE337'), ('Posto', 86.2, 'ISE337'),
        ('Kumir', 90.0, 'CS337'), ('Pagol', 59, 'CSE337'),
        ('Pagol', '60.0', 'CSE337'), ('Potol', 75, 'CSE337'),
        ('Potol', '93', 'ISE337'), ('Posto', 22, 'SE337')]
        f.write_sheets(data)
        self.assertEqual([], f.get_workbook_ave('^337$'))
